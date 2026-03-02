#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
独立测试：在无代理环境下用国内镜像安装 openpyxl，并验证可用。
运行: python3 test_openpyxl_install.py
"""
import os
import sys
import subprocess
from pathlib import Path

def main():
    mirror = "https://pypi.tuna.tsinghua.edu.cn/simple"
    trusted = "pypi.tuna.tsinghua.edu.cn"

    # 1. 尝试直接导入
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "测试"
        out = Path(__file__).parent / "test_openpyxl_out.xlsx"
        wb.save(out)
        print("OK: openpyxl 已安装，导入与写入正常")
        if out.exists():
            print(f"    已生成: {out}")
        return 0
    except ImportError as e:
        print(f"未安装: {e}")
    except Exception as e:
        print(f"导入/写入异常: {e}")
        return 1

    # 2. 无代理 + 国内镜像安装
    print(f"正在安装 openpyxl（禁用代理，镜像: {mirror}）...")
    env = os.environ.copy()
    for k in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy", "NO_PROXY", "no_proxy"):
        env.pop(k, None)
    cmd = [
        sys.executable, "-m", "pip", "install", "openpyxl",
        "-i", mirror, "--trusted-host", trusted,
        "--disable-pip-version-check",
    ]
    try:
        r = subprocess.run(cmd, env=env, timeout=120, capture_output=True, text=True)
        if r.returncode != 0:
            print("安装失败 stdout:", r.stdout or "(无)")
            print("安装失败 stderr:", r.stderr or "(无)")
            return 1
        print("安装命令执行完成。")
    except subprocess.TimeoutExpired:
        print("安装超时（120s）")
        return 1
    except Exception as e:
        print(f"安装过程异常: {e}")
        return 1

    # 3. 再次验证
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "测试"
        out = Path(__file__).parent / "test_openpyxl_out.xlsx"
        wb.save(out)
        print("OK: 安装后 openpyxl 可用，已生成 test_openpyxl_out.xlsx")
        return 0
    except Exception as e:
        print(f"安装后仍失败: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main())
