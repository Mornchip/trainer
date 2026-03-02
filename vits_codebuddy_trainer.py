#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
VITS CodeBuddy Trainer - 集成音频数据清洗功能
包含：数据清洗、Excel报告导出、VITS训练
"""

import os
import re
import sys
import time
import json
import shutil
import argparse
import subprocess
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# 禁用 Coqui 训练统计上报，避免连 coqui.gateway.scarf.sh 时的代理/网络错误
os.environ["TRAINER_TELEMETRY"] = "0"
# 避免 trainer 探测 git 时刷屏 "fatal: not a git repository"
os.environ["GIT_TERMINAL_PROMPT"] = "0"

# 依赖: pip install openpyxl pandas torch torchaudio
try:
    from audio_align_repair import run_repair
except ImportError:
    run_repair = None
    print("⚠️  audio_align_repair 模块未找到，跳过对齐修复步骤")
# 脚本所在目录，用于默认数据路径（从任意工作目录运行都能找到 segments）
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_PATH = str(SCRIPT_DIR / "segments")
DEFAULT_META_FILE = "meta_train_vits.txt"
LOOPMAX=15000
# build_dataset_from_labels 使用的路径与规模常量（本文件自包含，不依赖其他脚本）
ROOT = SCRIPT_DIR
AUDIO_DIR = "segments"
DATASET_LABELS = SCRIPT_DIR.parent / "dataset" / "labels"
DATASET_DOWNLOADS = SCRIPT_DIR.parent / "dataset" / "downloads"
TRAIN_SAMPLES_TARGET = 500
VAL_SAMPLES_TARGET = 750
VAL_SAMPLES = 750

# 断点续训：可设置环境变量 VITS_ 指定续训目录，否则自动取 vits_codebuddy_output 下最新 run
VITS_OUTPUT_DIR = SCRIPT_DIR / "vits_codebuddy_output"


def _is_tts_data_ready(path):
    """判断目录下是否已有 TTS 所需 train.csv、val.csv、meta_train_vits.txt，无需再准备数据。"""
    p = Path(path)
    return (
        (p / "train.csv").exists()
        and (p / "val.csv").exists()
        and (p / "meta_train_vits.txt").exists()
    )


def _get_train_data_path(base_path):
    """返回可直接用于训练的 data_path：优先 cleaned_audio，否则 base_path（若已具备 TTS 数据）。"""
    base = Path(base_path).resolve()
    if _is_tts_data_ready(base):
        return base
    cleaned = base / "cleaned_audio"
    if cleaned.is_dir() and _is_tts_data_ready(cleaned):
        return cleaned
    return None


def _get_latest_run_dir():
    """在 vits_codebuddy_output 下找含 config.json 且含 checkpoint 的 run 目录（修改时间最新），用于断点续训；无 checkpoint 则返回空。"""
    if not VITS_OUTPUT_DIR.is_dir():
        return ""
    best_dir, best_mtime = None, 0.0
    for d in VITS_OUTPUT_DIR.iterdir():
        if not d.is_dir():
            continue
        if not (d / "config.json").exists():
            continue
        ckpts = list(d.glob("best_model_*.pth")) + list(d.glob("checkpoint_*.pth"))
        if not ckpts:
            continue
        mtime = max(p.stat().st_mtime for p in ckpts)
        if mtime > best_mtime:
            best_mtime = mtime
            best_dir = d
    if best_dir is None:
        return ""
    return str(best_dir)


# 用于每轮计时：在 EVAL PERFORMANCE 块后注入「本轮时间 | 加载数据 | 累计时间」
_vits_epoch_timing = {"train_start": None, "last_epoch_end": None}


def _format_loss_decimals(text):
    """将行内浮点数格式化为保留 2 位小数（用于训练损失等指标显示）。"""
    import re
    def repl(m):
        try:
            return f"{float(m.group(0)):.2f}"
        except ValueError:
            return m.group(0)
    return re.sub(r"\d+\.\d+", repl, text)


class LossFormatStream:
    """包装流：对包含 loss/epoch 的行将浮点数格式化为 2 位小数后输出。"""
    def __init__(self, underlying):
        self._out = underlying
        self._buf = ""

    def write(self, data):
        self._buf += data
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            line_lower = line.lower()
            if "loss" in line_lower or "epoch" in line_lower or "step" in line_lower:
                line = _format_loss_decimals(line)
            self._out.write(line + "\n")

    def flush(self):
        if self._buf:
            line_lower = self._buf.lower()
            if "loss" in line_lower or "epoch" in line_lower or "step" in line_lower:
                self._buf = _format_loss_decimals(self._buf)
            self._out.write(self._buf)
            self._buf = ""
        self._out.flush()

    def close(self):
        self.flush()
        if hasattr(self._out, "close"):
            self._out.close()


class EpochTimingStream:
    """在每轮 EVAL PERFORMANCE 块后注入：本轮训练时间、加载数据时间、累计时间。"""
    def __init__(self, underlying):
        self._out = underlying
        self._buf = ""
        self._in_eval_block = False
        self._loader_time = None
        self._step_time = None
        self._epoch_dur = None
        self._total_dur = None

    def _strip_ansi(self, text):
        return re.sub(r"\x1b\[[0-9;]*m", "", text)

    def _write_timing_line(self):
        if self._epoch_dur is None and self._total_dur is None:
            return
        epoch_s = self._epoch_dur if self._epoch_dur is not None else 0.0
        total_s = self._total_dur if self._total_dur is not None else 0.0
        total_min = total_s / 60.0
        loader_s = f"{self._loader_time:.3f}" if self._loader_time is not None else "—"
        step_s = f"{self._step_time:.2f}" if self._step_time is not None else "—"
        line = (
            f"  [Epoch 计时] 本轮训练: {epoch_s:.1f}s | "
            f"加载数据: {loader_s}s/step | step: {step_s}s | "
            f"累计: {total_min:.1f}min\n"
        )
        self._out.write(line)

    def write(self, data):
        self._buf += data
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            plain = self._strip_ansi(line)
            if "EVAL PERFORMANCE" in plain:
                self._in_eval_block = True
                self._loader_time = None
                self._step_time = None
                now = time.time()
                start = _vits_epoch_timing.get("train_start")
                last = _vits_epoch_timing.get("last_epoch_end")
                if start is not None:
                    self._total_dur = now - start
                    self._epoch_dur = (now - last) if last is not None else (now - start)
                _vits_epoch_timing["last_epoch_end"] = now
            if self._in_eval_block:
                if "avg_loader_time" in plain:
                    m = re.search(r"avg_loader_time[:\s]+([\d.]+)", plain)
                    if m:
                        try:
                            self._loader_time = float(m.group(1))
                        except ValueError:
                            pass
                if "avg_step_time" in plain:
                    m = re.search(r"avg_step_time[:\s]+([\d.]+)", plain)
                    if m:
                        try:
                            self._step_time = float(m.group(1))
                        except ValueError:
                            pass
            self._out.write(line + "\n")
            if self._in_eval_block and "avg_grad_norm_1" in plain:
                self._write_timing_line()
                self._in_eval_block = False
                self._epoch_dur = None
                self._total_dur = None

    def flush(self):
        if self._buf:
            self._out.write(self._buf)
            self._buf = ""
        self._out.flush()

    def close(self):
        self.flush()
        if hasattr(self._out, "close"):
            self._out.close()


class TeeLogger:
    """将 stdout 同时输出到控制台和日志文件，便于保存完整训练日志"""
    def __init__(self, console, log_path):
        self.console = console
        self.log_path = Path(log_path)
        self.file = open(self.log_path, "w", encoding="utf-8")

    def write(self, data):
        self.console.write(data)
        self.file.write(data)
        self.file.flush()

    def flush(self):
        self.console.flush()
        self.file.flush()

    def close(self):
        self.file.close()
        self.file = None

try:
    import torch
    import torch.nn as nn
    from torch.utils.data import Dataset, DataLoader
except ImportError:
    torch = None
    print("⚠️ PyTorch 未安装，训练功能不可用")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("⚠️ openpyxl 未安装，Excel导出功能不可用。运行: pip install openpyxl")


# ==================== 配置类 ====================
class CleanConfig:
    """
    音频清洗配置类。
    训练前会做：去静音、响度标准化、限幅/去削波、时长-文本合理过滤、采样率/声道统一、文本清洗。
    """
    # 时长限制（硬性规则）；呻吟语音建议 1s 起以保留更多短句
    MIN_DURATION = 1.0          # 低于此值直接丢弃（放宽至1秒，避免数据量不足）
    MAX_DURATION = 16.0          # 高于此值也丢弃
    
    # 音频质量检测阈值；呻吟语音建议放宽 SNR/静音 以回收更多样本
    SILENCE_THRESHOLD = 500     # 静音检测用阈值
    CLIPPING_THRESHOLD = 32000  # 削波检测阈值（超此视为爆音，结合下方 alimiter 处理）
    NOISE_FLOOR = 500           # 底噪阈值
    MIN_SNR_DB = 0.5            # 最小信噪比(dB)，低于此过滤（呻吟语音放宽至 0.5）
    MAX_SILENCE_RATIO_PERCENT = 85.0  # 静音比例上限(%)，超过此过滤（放宽至 85%）
    
    # 处理参数
    TARGET_DB = -16             # 目标响度(dB)
    MAX_SILENCE_DURATION = 0.3  # 最大允许静音段(秒)
    MIN_VOICE_RATIO = 0.3       # 最小语音比例
    # 限幅/去削波：消除尖波，输出峰值不超过此比例（1.0=不超满幅），避免爆音影响训练
    ALIMITER_LIMIT = 0.99       # FFmpeg alimiter limit
    # 时长-文本合理：每字合理时长范围(秒)，用于过滤“时长与文本明显不匹配”的未对齐片段
    DURATION_PER_CHAR_MIN = 0.12 # 每字最少约 0.12 秒
    DURATION_PER_CHAR_MAX = 1.8 # 每字最多约 0.35 秒
    ALIGN_STRICT_RATIO = 0.4    # 实际时长若低于 预期最小*此值 或 高于 预期最大/此值 则判为未对齐

    # 文本过滤参数
    MIN_CHINESE_CHARS = 1       # 最少中文字符数
    MAX_SAMPLES = 0             # 最大样本数(0表示不限制)
    
    # 路径配置
    CLEANED_SUBDIR = "cleaned_audio"
    BACKUP_SUBDIR = "backup_original"

    # 音频输出格式（与 TTS 训练一致）
    SAMPLE_RATE = 22050
    CHANNELS = 1

    # 底噪/电流声去除
    HIGHPASS_HZ = 80             # 高通滤波截止频率(Hz)，去除 50/60Hz 工频及低频底噪，0 表示关闭
    ENABLE_DENOISE = True        # 是否启用 FFT 频域降噪（进一步削弱底噪）
    AFFTDN_NF = -25              # FFT 降噪噪声底(dB)，越小降噪越强（-25 为适中）


def clean_text_for_tts(text):
    """
    文本清洗：统一标点、空格与特殊符号，便于 TTS 模型学习。
    - 去除首尾空白
    - 全角空格转半角，连续空白合并为单个空格
    - 去除控制字符、零宽字符
    - 中文标点保留（，。？！等），仅规范化空白
    """
    if not text or not isinstance(text, str):
        return "" if text is None else str(text).strip()
    # 去除控制字符、零宽字符
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    text = re.sub(r"[\u200b-\u200d\ufeff]", "", text)
    # 全角空格 -> 半角
    text = text.replace("\u3000", " ")
    # 首尾去空白
    text = text.strip()
    # 连续空白（含 \t \n）合并为单个空格
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ==================== 音频检测类 ====================
class AudioChecker:
    """音频质量检测类"""
    
    @staticmethod
    def check_duration(audio_path):
        """检测音频时长"""
        try:
            import wave
            with wave.open(str(audio_path), 'rb') as wf:
                frames = wf.getnframes()
                rate = wf.getframerate()
                duration = frames / float(rate)
                return duration
        except Exception as e:
            return 0
    
    @staticmethod
    def check_clipping(audio_path):
        """检测音频削波"""
        try:
            import wave
            import audioop
            with wave.open(str(audio_path), 'rb') as wf:
                frames = wf.readframes(wf.getnframes())
                if not frames:
                    return False
                max_val = max(abs(int.from_bytes(frames[i:i+2], 'little', signed=True)) 
                             for i in range(0, len(frames), 2) if i+2 <= len(frames))
                return max_val > CleanConfig.CLIPPING_THRESHOLD
        except:
            return False
    
    @staticmethod
    def check_silence(audio_path):
        """检测静音比例"""
        try:
            import wave
            with wave.open(str(audio_path), 'rb') as wf:
                frames = wf.readframes(wf.getnframes())
                if not frames:
                    return 1.0
                silent_frames = sum(1 for i in range(0, len(frames), 2) 
                                   if i+2 <= len(frames) and 
                                   abs(int.from_bytes(frames[i:i+2], 'little', signed=True)) < CleanConfig.SILENCE_THRESHOLD)
                return silent_frames / (len(frames)/2)
        except:
            return 0.5
    
    @staticmethod
    def estimate_snr(audio_path):
        """估算信噪比"""
        try:
            import wave
            import audioop
            with wave.open(str(audio_path), 'rb') as wf:
                frames = wf.readframes(wf.getnframes())
                if not frames:
                    return 0
                rms = audioop.rms(frames, 2)
                # 简化的SNR估算
                return 20 * (rms / 32767) if rms > 0 else 0
        except:
            return 0


# ==================== 音频处理类 ====================
class AudioProcessor:
    """音频处理类 - 使用FFmpeg"""
    
    @staticmethod
    def remove_silence_and_normalize(input_path, output_path):
        """
        去除静音、音量标准化、限幅去削波，并统一为 TTS 标准格式。
        - highpass: 去除低频电流声/底噪（50/60Hz 工频及以下）
        - afftdn: FFT 频域降噪，削弱底噪
        - silenceremove: 去掉首尾静音
        - loudnorm: 响度标准化
        - alimiter: 限幅/去削波，避免爆音失真
        - 输出: 22050Hz 单声道（与 CleanConfig.SAMPLE_RATE/CHANNELS 一致）
        """
        try:
            limit = getattr(CleanConfig, 'ALIMITER_LIMIT', 0.99)
            sr = getattr(CleanConfig, 'SAMPLE_RATE', 22050)
            ch = getattr(CleanConfig, 'CHANNELS', 1)
            highpass_hz = getattr(CleanConfig, 'HIGHPASS_HZ', 80)
            enable_denoise = getattr(CleanConfig, 'ENABLE_DENOISE', True)
            afftdn_nf = getattr(CleanConfig, 'AFFTDN_NF', -25)
            # 滤波器链：先降底噪/电流声，再去静音、响度、限幅
            af_parts = []
            if highpass_hz and highpass_hz > 0:
                af_parts.append(f'highpass=f={highpass_hz}')
            if enable_denoise:
                af_parts.append(f'afftdn=nf={afftdn_nf}')
            af_parts.extend([
                f'silenceremove=start_periods=1:start_duration=0.1:start_threshold=0.02:'
                f'stop_periods=-1:stop_duration=0.1:stop_threshold=0.02',
                f'loudnorm=I={CleanConfig.TARGET_DB}:TP=-1.5:LRA=11',
                f'alimiter=limit={limit}:attack=5:release=50:level=1'
            ])
            af_str = ','.join(af_parts)
            cmd = [
                'ffmpeg', '-y', '-hide_banner', '-loglevel', 'error',
                '-i', str(input_path),
                '-af', af_str,
                '-ar', str(sr), '-ac', str(ch),
                str(output_path)
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            return result.returncode == 0
        except Exception as e:
            print(f"  ⚠️ 处理失败: {e}")
            return False


# ==================== 数据清洗主类 ====================
class DataCleaner:
    """数据清洗主类 - 支持Excel报告导出"""

    def __init__(self, input_dir, meta_file):
        self.input_dir = Path(input_dir)
        self.meta_file = self.input_dir / meta_file
        self.output_dir = self.input_dir / CleanConfig.CLEANED_SUBDIR
        self.backup_dir = self.input_dir / CleanConfig.BACKUP_SUBDIR
        self.stats = defaultdict(int)
        self.detailed_records = []  # 用于Excel导出
        self.checker = AudioChecker()
        self.processor = AudioProcessor()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _count_chinese_chars(self, text):
        chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
        return len(chinese_pattern.findall(text))

    def _backup_original(self):
        if self.backup_dir.exists():
            return
        print(f"📦 备份原始数据到: {self.backup_dir}")
        self.backup_dir.mkdir(exist_ok=True)
        if self.meta_file.exists():
            shutil.copy2(self.meta_file, self.backup_dir / self.meta_file.name)

    def _load_metadata(self):
        metadata = []
        if not self.meta_file.exists():
            print(f"⚠️ 标注文件不存在: {self.meta_file}")
            return metadata
        with open(self.meta_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = [p.strip() for p in line.split('|')]
                if len(parts) >= 2:
                    # 支持 path|text 或 path|speaker|text；文本做统一清洗后用于训练
                    audio_file = parts[0]
                    text = clean_text_for_tts(parts[-1])
                    if not audio_file.lower().endswith('.wav'):
                        audio_file = audio_file + '.wav'
                    metadata.append({'audio_file': audio_file, 'text': text})
        return metadata

    def _check_audio_quality_with_details(self, item):
        audio_path = self.input_dir / item['audio_file']
        text = item.get('text', '')
        record = {
            'filename': item['audio_file'], 'text': text, 'chinese_chars': self._count_chinese_chars(text),
            'duration': 0, 'status': 'unknown', 'fail_reason': '', 'clipping_detected': False,
            'silence_ratio': 0, 'snr_db': 0, 'processed': False, 'output_file': ''
        }
        if not audio_path.exists():
            record['status'] = 'failed'
            record['fail_reason'] = '文件不存在'
            self.stats['file_not_found'] += 1
            return None, record
        duration = self.checker.check_duration(audio_path)
        record['duration'] = round(duration, 2)
        if duration < CleanConfig.MIN_DURATION:
            self.stats['too_short'] += 1
            record['status'] = 'filtered'
            record['fail_reason'] = f'时长过短({duration:.2f}s < {CleanConfig.MIN_DURATION}s)'
            return None, record
        if duration > CleanConfig.MAX_DURATION:
            self.stats['too_long'] += 1
            record['status'] = 'filtered'
            record['fail_reason'] = f'时长过长({duration:.2f}s > {CleanConfig.MAX_DURATION}s)'
            return None, record
        if CleanConfig.MIN_CHINESE_CHARS > 0:
            chinese_count = record['chinese_chars']
            if chinese_count < CleanConfig.MIN_CHINESE_CHARS:
                self.stats['skipped_chinese'] += 1
                record['status'] = 'filtered'
                record['fail_reason'] = f'中文字符不足({chinese_count} < {CleanConfig.MIN_CHINESE_CHARS})'
                return None, record
        # 时长-文本对齐检查：过滤“时长与字数明显不匹配”的疑似未对齐片段
        chinese_count = record['chinese_chars']
        if chinese_count > 0:
            expected_min = chinese_count * CleanConfig.DURATION_PER_CHAR_MIN
            expected_max = chinese_count * CleanConfig.DURATION_PER_CHAR_MAX
            r = CleanConfig.ALIGN_STRICT_RATIO
            if duration < expected_min * r or duration > expected_max / r:
                self.stats['misaligned'] = self.stats.get('misaligned', 0) + 1
                record['status'] = 'filtered'
                record['fail_reason'] = (
                    f'时长与文本长度不匹配(疑似未对齐): {duration:.2f}s, '
                    f'约{chinese_count}字预期{expected_min:.2f}-{expected_max:.2f}s'
                )
                return None, record
        record['clipping_detected'] = self.checker.check_clipping(audio_path)
        if record['clipping_detected']:
            self.stats['clipping'] += 1
        record['silence_ratio'] = round(self.checker.check_silence(audio_path) * 100, 2)
        record['snr_db'] = round(self.checker.estimate_snr(audio_path), 2)
        record['status'] = 'passed'
        return item, record

    def _process_audio_with_record(self, item, record):
        input_path = self.input_dir / item['audio_file']
        output_path = self.output_dir / Path(item['audio_file']).name
        success = self.processor.remove_silence_and_normalize(input_path, output_path)
        if success and output_path.exists():
            record['processed'] = True
            record['output_file'] = str(output_path.relative_to(self.input_dir))
            new_duration = self.checker.check_duration(output_path)
            record['duration'] = round(new_duration, 2)
            return {'audio_file': str(output_path.relative_to(self.input_dir)), 'text': item['text']}
        self.stats['processing_failed'] += 1
        record['status'] = 'failed'
        record['fail_reason'] = 'FFmpeg处理失败'
        return None

    def _limit_samples(self, items):
        if CleanConfig.MAX_SAMPLES <= 0 or len(items) <= CleanConfig.MAX_SAMPLES:
            return items
        import random
        random.seed(42)
        return random.sample(items, CleanConfig.MAX_SAMPLES)

    def _generate_cleaned_metadata(self, cleaned_items):
        output_meta = self.output_dir / "cleaned_metadata.txt"
        with open(output_meta, 'w', encoding='utf-8') as f:
            for item in cleaned_items:
                f.write(f"{item['audio_file']}|{item['text']}\n")
        return output_meta

    def _generate_excel_report(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # 使用国内镜像尝试安装（子进程内禁用代理，避免 ProxyError）
            mirror = "https://pypi.tuna.tsinghua.edu.cn/simple"
            print(f"⚠️ openpyxl 未安装，正在使用国内镜像安装: {mirror}")
            try:
                env = os.environ.copy()
                for key in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy", "ALL_PROXY", "all_proxy"):
                    env.pop(key, None)
                subprocess.run(
                    [sys.executable, "-m", "pip", "install", "openpyxl", "-i", mirror, "-q", "--trusted-host", "pypi.tuna.tsinghua.edu.cn"],
                    check=True,
                    timeout=120,
                    env=env,
                )
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                print("✅ openpyxl 安装成功，继续生成 Excel 报告")
            except (subprocess.CalledProcessError, subprocess.TimeoutExpired, ImportError) as e:
                print(f"⚠️ 安装失败或仍无法导入，跳过 Excel 报告: {e}")
                return None
        excel_path = self.input_dir / f"cleaning_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "处理摘要"
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        ws_summary['A1'] = "🧹 音频数据清洗报告"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:E1')
        ws_summary['A3'] = "生成时间:"
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_summary['A4'] = "输入目录:"
        ws_summary['B4'] = str(self.input_dir)
        ws_summary['A5'] = "输出目录:"
        ws_summary['B5'] = str(self.output_dir)
        ws_summary['A7'] = "📊 统计摘要"
        ws_summary['A7'].font = title_font
        summary_headers = ['项目', '数量', '占比']
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        total = self.stats['total']
        kept = len([r for r in self.detailed_records if r['processed']])
        filtered = len([r for r in self.detailed_records if r['status'] == 'filtered'])
        failed = len([r for r in self.detailed_records if r['status'] == 'failed'])
        summary_data = [
            ['总样本数', total, '100%'],
            ['保留样本', kept, f'{kept/total*100:.1f}%' if total > 0 else '0%'],
            ['过滤样本', filtered, f'{filtered/total*100:.1f}%' if total > 0 else '0%'],
            ['处理失败', failed, f'{failed/total*100:.1f}%' if total > 0 else '0%'],
            ['', '', ''], ['【过滤原因详情】', '', ''],
            ['时长过短(<1.5s)', self.stats['too_short'], ''],
            ['时长过长(>7.0s)', self.stats['too_long'], ''],
            ['中文不足', self.stats['skipped_chinese'], ''],
            ['时长与文本未对齐', self.stats.get('misaligned', 0), ''],
            ['文件不存在', self.stats['file_not_found'], ''],
            ['处理失败', self.stats['processing_failed'], ''],
        ]
        for row_idx, row_data in enumerate(summary_data, start=9):
            for col_idx, value in enumerate(row_data, start=1):
                ws_summary.cell(row=row_idx, column=col_idx, value=value)
        ws_summary['A22'] = "⚙️ 配置参数"
        ws_summary['A22'].font = title_font
        config_data = [
            ['最小时长', f'{CleanConfig.MIN_DURATION}s'], ['最大时长', f'{CleanConfig.MAX_DURATION}s'],
            ['最少中文字符', CleanConfig.MIN_CHINESE_CHARS],
            ['最大样本数', CleanConfig.MAX_SAMPLES if CleanConfig.MAX_SAMPLES > 0 else '无限制'],
            ['目标响度', f'{CleanConfig.TARGET_DB}dB'], ['削波阈值', CleanConfig.CLIPPING_THRESHOLD],
            ['静音阈值', CleanConfig.SILENCE_THRESHOLD],
            ['限幅(消除尖波)', f'alimiter limit={getattr(CleanConfig, "ALIMITER_LIMIT", 0.99)}'],
            ['高通(去电流声/底噪)', f'{getattr(CleanConfig, "HIGHPASS_HZ", 80)}Hz'],
            ['FFT降噪(底噪)', '开启' if getattr(CleanConfig, 'ENABLE_DENOISE', True) else '关闭'],
            ['时长-文本对齐(每字)', f'{getattr(CleanConfig, "DURATION_PER_CHAR_MIN", 0.12)}~{getattr(CleanConfig, "DURATION_PER_CHAR_MAX", 0.35)}s'],
        ]
        for row_idx, (key, value) in enumerate(config_data, start=23):
            ws_summary.cell(row=row_idx, column=1, value=key)
            ws_summary.cell(row=row_idx, column=2, value=value)
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 10
        ws_detail = wb.create_sheet("详细记录")
        df = pd.DataFrame(self.detailed_records)
        column_mapping = {
            'filename': '文件名', 'text': '文本内容', 'chinese_chars': '中文字符数', 'duration': '时长(秒)',
            'status': '处理状态', 'fail_reason': '失败原因', 'clipping_detected': '削波检测',
            'silence_ratio': '静音比例(%)', 'snr_db': '信噪比(dB)', 'processed': '是否处理', 'output_file': '输出文件'
        }
        df = df[[col for col in column_mapping.keys() if col in df.columns]]
        df.rename(columns=column_mapping, inplace=True)
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws_detail.cell(row=r_idx, column=c_idx, value=value)
        for c_idx, header in enumerate(df.columns, 1):
            cell = ws_detail.cell(row=1, column=c_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for column in ws_detail.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws_detail.column_dimensions[column_letter].width = min(max_length + 2, 60)
        ws_kept = wb.create_sheet("保留样本")
        kept_records = [r for r in self.detailed_records if r['processed']]
        if kept_records:
            df_kept = pd.DataFrame(kept_records)
            df_kept = df_kept[['filename', 'text', 'chinese_chars', 'duration', 'output_file']]
            df_kept.rename(columns={'filename': '文件名', 'text': '文本内容', 'chinese_chars': '中文字符数', 'duration': '时长(秒)', 'output_file': '输出路径'}, inplace=True)
            for r_idx, row in enumerate(df_kept.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws_kept.cell(row=r_idx, column=c_idx, value=value)
            for c_idx, header in enumerate(df_kept.columns, 1):
                cell = ws_kept.cell(row=1, column=c_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
        for column in ws_kept.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws_kept.column_dimensions[column_letter].width = min(max_length + 2, 60)
        # 失败/过滤记录：仅保存未通过或处理失败的记录
        failed_records = [r for r in self.detailed_records if r.get('status') in ('filtered', 'failed') or not r.get('processed', False)]
        ws_failed = wb.create_sheet("失败与过滤记录")
        if failed_records:
            df_failed = pd.DataFrame(failed_records)
            fail_columns = ['filename', 'text', 'chinese_chars', 'duration', 'status', 'fail_reason', 'clipping_detected', 'silence_ratio', 'snr_db']
            df_failed = df_failed[[c for c in fail_columns if c in df_failed.columns]]
            df_failed.rename(columns={
                'filename': '文件名', 'text': '文本内容', 'chinese_chars': '中文字符数', 'duration': '时长(秒)',
                'status': '处理状态', 'fail_reason': '失败/过滤原因', 'clipping_detected': '削波检测',
                'silence_ratio': '静音比例(%)', 'snr_db': '信噪比(dB)'
            }, inplace=True)
            for r_idx, row in enumerate(df_failed.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws_failed.cell(row=r_idx, column=c_idx, value=value)
            for c_idx, header in enumerate(df_failed.columns, 1):
                cell = ws_failed.cell(row=1, column=c_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            for column in ws_failed.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws_failed.column_dimensions[column_letter].width = min(max_length + 2, 60)
        else:
            ws_failed.cell(row=1, column=1, value="无失败或过滤记录")
        wb.save(excel_path)
        print(f"📊 Excel报告已生成: {excel_path}")
        return str(excel_path)

    def run(self):
        print("=" * 70)
        print("🧹 开始数据清洗流程")
        print("=" * 70)
        self._backup_original()
        metadata = self._load_metadata()
        self.stats['total'] = len(metadata)
        print(f"📁 加载了 {len(metadata)} 条标注记录")
        passed_items = []
        intloop=0
        for item in metadata:
            if intloop>LOOPMAX:
                break
            intloop+=1
            result, record = self._check_audio_quality_with_details(item)
            print(record)
            self.detailed_records.append(record)
            if result:
                passed_items.append((result, record))
            else:
                self.stats['failed'] += 1
        print(f"🔍 质量筛选: {len(passed_items)}/{len(metadata)} 通过初步筛选")
        passed_items = self._limit_samples(passed_items)
        cleaned_items = []
        intloop=0
        for item, record in passed_items:
            if intloop>LOOPMAX:
                break
            intloop+=1
            result = self._process_audio_with_record(item, record)
            if result:
                cleaned_items.append(result)
        cleaned_meta_path = self._generate_cleaned_metadata(cleaned_items)
        excel_path = self._generate_excel_report()
        print("=" * 70)
        print(f"✅ 清洗完成: {len(cleaned_items)}/{len(metadata)} 保留")
        print(f"📄 清洗后标注: {cleaned_meta_path}")
        print("=" * 70)
        return {
            'output_dir': str(self.output_dir),
            'cleaned_meta': str(cleaned_meta_path),
            'excel_report': excel_path,
            'total': len(metadata),
            'kept': len(cleaned_items)
        }


# ==================== 从 labels 构建数据集（本文件自包含，不调用其他代码文件）====================
def _label_count_chinese_chars(text):
    """标注中汉字个数，用于 build_dataset_from_labels 过滤。"""
    if not text or not isinstance(text, str):
        return 0
    return len(re.findall(r"[\u4e00-\u9fff]", text))


def _find_mp3_for_basename(basename, downloads_dir):
    """根据标注文件名找 downloads 下同名音频。支持无空格 stem 与「数字 空格 其余」变体。"""
    downloads_dir = Path(downloads_dir)
    candidates = [basename]
    if basename and basename[0].isdigit():
        i = 0
        while i < len(basename) and basename[i].isdigit():
            i += 1
        if i < len(basename):
            candidates.append(basename[:i] + " " + basename[i:])
    for stem in candidates:
        for ext in (".mp3", ".wav", ".flac", ".m4a"):
            p = downloads_dir / (stem + ext)
            if p.exists():
                return p
    return None


def _parse_label_line(line):
    """解析一行标注：basename start_sec end_sec flag text。返回 (basename, start, end, flag, text) 或 None。"""
    line = line.strip()
    if not line:
        return None
    parts = line.split(None, 4)
    separator = "--小烟"
    if separator in line:
        end_index = line.find(separator) + len(separator)
        file_name = line[:end_index].strip()
        line = line.replace(file_name, file_name + " ")
        parts = line.split(None, 4)
    if len(parts) < 5:
        return None
    try:
        if len(parts) == 5:
            base = parts[0]
            start = float(parts[1].replace(",", ""))
            end = float(parts[2].replace(",", ""))
            speaker = parts[3]
            text = parts[4].strip()
        elif len(parts) == 4:
            base, start, end = parts[0], float(parts[1].replace(",", "")), float(parts[2].replace(",", ""))
            text = parts[3].strip()
            speaker = 0
        else:
            return None
    except (ValueError, IndexError):
        return None
    if not text or end <= start or end - start < 0.01:
        return None
    return (base, start, end, speaker, text)


def _check_data_quality():
    """检查原始标注数据质量，返回 (是否通过, 统计信息)。"""
    print("\n" + "=" * 60)
    print("📊 数据质量预检查")
    print("=" * 60)
    if not DATASET_LABELS.is_dir():
        print(f"❌ 标注目录不存在: {DATASET_LABELS}")
        return False, {}
    all_texts = []
    all_durations = []
    total_files = 0
    for label_file in sorted(DATASET_LABELS.glob("*.txt")):
        total_files += 1
        with open(label_file, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                parsed = _parse_label_line(line)
                if parsed:
                    base, start_sec, end_sec, speaker, text = parsed
                    all_texts.append(text)
                    all_durations.append(end_sec - start_sec)
    if not all_texts:
        print("❌ 没有找到任何有效标注数据")
        return False, {}
    lengths = [_label_count_chinese_chars(t) for t in all_texts]
    unique_texts = len(set(all_texts))
    total_texts = len(all_texts)
    diversity_ratio = unique_texts / total_texts
    from collections import Counter
    text_counts = Counter(all_texts)
    most_common = text_counts.most_common(5)
    print(f"\n📈 数据统计:")
    print(f"   标注文件数: {total_files}")
    print(f"   总样本数: {total_texts}")
    print(f"   唯一文本数: {unique_texts}")
    print(f"   多样性比例: {diversity_ratio:.1%}")
    print(f"\n📏 文本长度分布（中文字符）:")
    for i in range(1, 11):
        count = sum(1 for l in lengths if l == i)
        if count > 0:
            print(f"   {i}字: {count}条 ({count/len(lengths)*100:.1f}%)")
    long_count = sum(1 for l in lengths if l > 10)
    if long_count > 0:
        print(f"   >10字: {long_count}条")
    print(f"\n⏱️ 时长分布:")
    short_dur = sum(1 for d in all_durations if d < 1.5)
    mid_dur = sum(1 for d in all_durations if 1.5 <= d <= 7.0)
    long_dur = sum(1 for d in all_durations if d > 7.0)
    print(f"   <1.5s: {short_dur}条 ({short_dur/len(all_durations)*100:.1f}%)")
    print(f"   1.5-7s: {mid_dur}条 ({mid_dur/len(all_durations)*100:.1f}%)")
    print(f"   >7s: {long_dur}条 ({long_dur/len(all_durations)*100:.1f}%)")
    print(f"\n🔥 重复最多的文本 (Top 5):")
    for text, count in most_common:
        print(f"   \"{text}\" 出现 {count} 次")
    issues = []
    if diversity_ratio < 0.3:
        issues.append(f"多样性比例过低 ({diversity_ratio:.1%} < 30%)")
    if most_common and most_common[0][1] > total_texts * 0.1:
        issues.append(f"文本重复严重 (\"{most_common[0][0]}\" 出现 {most_common[0][1]} 次)")
    if sum(1 for l in lengths if l >= 1) < total_texts * 0.5:
        issues.append("有效文本(>=1字)占比过低")
    print("\n" + "=" * 60)
    if issues:
        print("⚠️ 数据质量问题:")
        for issue in issues:
            print(f"   - {issue}")
        print("=" * 60)
        return False, {"total": total_texts, "unique": unique_texts, "diversity": diversity_ratio, "issues": issues}
    print("✅ 数据质量检查通过")
    print("=" * 60)
    return True, {"total": total_texts, "unique": unique_texts, "diversity": diversity_ratio}


def _clean_moaning_text_for_labels(text):
    """清理标注文本：仅做基本清理，保证文本-音频对齐。用于 build_dataset_from_labels。"""
    if not isinstance(text, str):
        text = str(text)
    text = text.replace("|", " ")
    text = re.sub(r'\s+', ' ', text).strip()
    return text if text else "啊"


def build_dataset_from_labels(min_chinese=1, min_duration=1.5, max_duration=7.0, sample_rate=22050, max_samples=0, skip_quality_check=False):
    """
    从 dataset/labels + dataset/downloads 构建训练数据：
    - 只保留标注长度 >= min_chinese 个汉字的样本
    - 只保留时长在 [min_duration, max_duration] 秒的片段
    - max_samples>0 时只保留前 max_samples 条（用于快速试训，如 1000 条）
    - 从 mp3 切条为 wav 写入 lasttraincodebuddy/segments，并生成 train.csv / val.csv
    - 自动进行数据质量检查，不合格会停止训练
    """
    import random
    import shutil
    try:
        import librosa
    except ImportError:
        print("❌ 需要 librosa，请安装: pip install librosa")
        return False
    try:
        import soundfile as sf
    except ImportError:
        sf = None
    try:
        from scipy.io import wavfile as scipy_wavfile
        from scipy.signal import butter, filtfilt
    except ImportError:
        scipy_wavfile = None
        butter = filtfilt = None

    if not DATASET_LABELS.is_dir():
        print(f"❌ 标注目录不存在: {DATASET_LABELS}")
        return False
    if not DATASET_DOWNLOADS.is_dir():
        print(f"❌ 音频目录不存在: {DATASET_DOWNLOADS}")
        return False
    
    # 数据质量预检查
    if not skip_quality_check:
        passed, stats = _check_data_quality()
        if not passed:
            print("\n❌ 数据质量检查未通过，停止训练！")
            print("请检查标注数据，解决上述问题后再训练。")
            print("如果确认要跳过检查，设置 skip_quality_check=True")
            return False

    segments_dir = ROOT / AUDIO_DIR
    if segments_dir.exists():
        if segments_dir.is_symlink():
            segments_dir.unlink()
        else:
            try:
                shutil.rmtree(segments_dir)
            except Exception as e:
                print(f"⚠️ 无法清空 {segments_dir}: {e}")
                return False
    segments_dir.mkdir(parents=True, exist_ok=True)

    rows = []
    skipped_short_text = 0
    skipped_duration = 0
    skipped_no_audio = 0
    skipped_error = 0
    skipped_duplicate = 0  # 新增：因重复过多被跳过
    text_occurrence = {}   # 新增：记录每个文本出现次数
    
    # 实时过滤参数
    MAX_TEXT_OCCURRENCE = 50  # 单个文本最多保留50次，超过的丢弃
    MIN_TEXT_LENGTH = 1       # 最少1个汉字
    MAX_TEXT_LENGTH = 50      # 最多50个汉字
    intloop=0
    for label_file in sorted(DATASET_LABELS.glob("*.txt")):
        if intloop>LOOPMAX:
            break
        stem = label_file.stem
        mp3_path = _find_mp3_for_basename(stem, DATASET_DOWNLOADS)
        if not mp3_path or not mp3_path.exists():
            skipped_no_audio += 1
            continue
        try:
            full_audio, sr = librosa.load(str(mp3_path), sr=sample_rate, mono=True)
        except Exception as e:
            skipped_error += 1
            continue
        with open(label_file, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                if intloop>LOOPMAX:
                    break
                intloop+=1
                parsed = _parse_label_line(line)
                if not parsed:
                    continue
                base, start_sec, end_sec, flag, text = parsed
                if _label_count_chinese_chars(text) < min_chinese:
                    skipped_short_text += 1
                    continue
                dur = end_sec - start_sec
                if dur < min_duration or dur > max_duration:
                    skipped_duration += 1
                    continue
                start_samp = int(start_sec * sr)
                end_samp = int(end_sec * sr)
                if start_samp >= len(full_audio) or end_samp <= start_samp:
                    skipped_error += 1
                    continue
                print(mp3_path.name,start_samp, end_samp,text,dur)
                segment_wav = full_audio[start_samp:end_samp]
                # 底噪/电流声：高通滤波去除低频（与 CleanConfig.HIGHPASS_HZ 一致）
                highpass_hz = getattr(CleanConfig, 'HIGHPASS_HZ', 80)
                if highpass_hz and highpass_hz > 0 and butter is not None and filtfilt is not None:
                    nyq = 0.5 * sample_rate
                    hp = min(highpass_hz, nyq - 1)
                    if hp > 0:
                        b, a = butter(4, hp / nyq, btype='high')
                        segment_wav = filtfilt(b, a, segment_wav)
                if len(segment_wav) < 0.2 * sr:
                    skipped_duration += 1
                    continue
                seg_name = f"segment_{len(rows):06d}"
                wav_path = segments_dir / f"{seg_name}.wav"
                try:
                    if sf:
                        sf.write(str(wav_path), segment_wav, sample_rate)
                    elif scipy_wavfile:
                        import numpy as _np
                        wav_int = (_np.clip(segment_wav, -1, 1) * 32767).astype(_np.int16)
                        scipy_wavfile.write(str(wav_path), sample_rate, wav_int)
                    else:
                        skipped_error += 1
                        continue
                except Exception:
                    skipped_error += 1
                    continue
                text_clean = _clean_moaning_text_for_labels(text)
                text_len = _label_count_chinese_chars(text_clean)
                
                # 实时过滤1：文本长度检查
                if not text_clean or text_len < MIN_TEXT_LENGTH or text_len > MAX_TEXT_LENGTH:
                    wav_path.unlink(missing_ok=True)
                    skipped_short_text += 1
                    continue
                
                # 实时过滤2：重复文本限制（单个文本最多保留MAX_TEXT_OCCURRENCE次）
                text_occurrence[text_clean] = text_occurrence.get(text_clean, 0) + 1
                if text_occurrence[text_clean] > MAX_TEXT_OCCURRENCE:
                    wav_path.unlink(missing_ok=True)
                    skipped_duplicate += 1
                    continue
                
                rows.append({"audio_path": f"{AUDIO_DIR}/{seg_name}.wav", "text": text_clean})
                if max_samples > 0 and len(rows) >= max_samples:
                    break
        print(f"   已处理 {len(rows)} 条...")
        if max_samples > 0 and len(rows) >= max_samples:
            break
        if len(rows) % 500 == 0 and len(rows) > 0:
            print(f"   已处理 {len(rows)} 条...")

    if not rows:
        print(f"❌ 从 labels 未得到任何合格样本（汉字>={min_chinese}，时长 {min_duration}-{max_duration}s）")
        print(f"   跳过: 标注<{min_chinese}汉字={skipped_short_text}, 时长不符={skipped_duration}, 重复过多={skipped_duplicate}, 无音频/错误={skipped_no_audio}/{skipped_error}")
        return False

    random.seed(42)
    random.shuffle(rows)
    if max_samples > 0 and len(rows) > max_samples:
        rows = rows[:max_samples]
        print(f"   已截断为 {max_samples} 条。")
    # 如果数据够1100条，严格按1000训练+100验证分割；不够则按比例
    total_needed = TRAIN_SAMPLES_TARGET + VAL_SAMPLES_TARGET
    if len(rows) >= total_needed:
        n_train = TRAIN_SAMPLES_TARGET
        n_val = VAL_SAMPLES_TARGET
    else:
        # 数据不足时，验证集至少10条，最多100条，其余给训练集
        n_val = min(VAL_SAMPLES_TARGET, max(10, len(rows) // 11))
        n_train = len(rows) - n_val

    train_rows = rows[:n_train]
    val_rows = rows[n_train:n_train + n_val]

    train_df = pd.DataFrame(train_rows)
    val_df = pd.DataFrame(val_rows)
    train_csv = ROOT / "train.csv"
    val_csv = ROOT / "val.csv"
    train_df.to_csv(train_csv, index=False, encoding="utf-8-sig")
    val_df.to_csv(val_csv, index=False, encoding="utf-8-sig")

    # 计算最终数据多样性
    unique_texts_final = len(set(r['text'] for r in rows))
    diversity_final = unique_texts_final / len(rows) if rows else 0
    
    print(f"✅ 已从 labels+downloads 生成: train={n_train}, val={n_val}（验证集=VAL_SAMPLES={VAL_SAMPLES}，汉字>={min_chinese}，时长 {min_duration}-{max_duration}s）")
    print(f"   跳过: 标注<{min_chinese}汉字={skipped_short_text}, 时长不符={skipped_duration}, 重复过多={skipped_duplicate}, 无音频={skipped_no_audio}, 错误={skipped_error}")
    print(f"   数据多样性: {unique_texts_final}/{len(rows)} ({diversity_final:.1%}) 唯一文本")
    print(f"   音频已写入 {segments_dir}，CSV 已写入 {train_csv} / {val_csv}")
    
    # 如果多样性仍然过低，给出警告
    if diversity_final < 0.3:
        print(f"\n⚠️ 警告: 数据多样性过低 ({diversity_final:.1%} < 30%)，建议检查原始标注数据")
    return True


# ==================== VITS数据集类 ====================
class VITSDataset(Dataset):
    """VITS数据集"""
    def __init__(self, data_path, meta_file):
        self.data_path = Path(data_path)
        self.meta_file = self.data_path / meta_file
        self.samples = []
        
        if self.meta_file.exists():
            with open(self.meta_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    parts = [p.strip() for p in line.split('|')]
                    if len(parts) >= 2:
                        audio = parts[0]
                        text = clean_text_for_tts(parts[-1])
                        if not audio.lower().endswith('.wav'):
                            audio = audio + '.wav'
                        self.samples.append({'audio': audio, 'text': text})
    
    def __len__(self):
        return len(self.samples)
    
    def __getitem__(self, idx):
        sample = self.samples[idx]
        audio_path = self.data_path / sample['audio']
        
        # 这里添加音频加载和特征提取逻辑
        # 返回 audio_features, text_tokens
        return {
            'audio_path': str(audio_path),
            'text': sample['text']
        }


# ==================== 自包含 VITS 训练（不依赖 vits_moaning_trainer.py）====================
def _prepare_data_tts_for_train(root_path, audio_dir="."):
    """为 TTS 准备 meta 与 wavs 链接，返回 (meta_train, meta_val, num_samples, chars_from_data)。"""
    root_path = Path(root_path).resolve()
    train_csv = root_path / "train.csv"
    val_csv = root_path / "val.csv"
    if not train_csv.exists() or not val_csv.exists():
        return None, None, 0, ""
    train_df = pd.read_csv(train_csv, dtype={"text": str, "audio_path": str})
    val_df = pd.read_csv(val_csv, dtype={"text": str, "audio_path": str})
    segments_dir = root_path / audio_dir
    if not segments_dir.exists():
        segments_dir.mkdir(parents=True, exist_ok=True)
    # 计算 duration 并过滤：1.5–7s，语速>=1.5字/秒
    try:
        import librosa
        def get_dur(path):
            try:
                return librosa.get_duration(path=str(root_path / path))
            except Exception:
                return 0.0
        train_df["duration"] = train_df["audio_path"].apply(get_dur)
        val_df["duration"] = val_df["audio_path"].apply(get_dur)
    except ImportError:
        pass
    if "duration" in train_df.columns:
        train_df = train_df[(train_df["duration"] >= 1.5) & (train_df["duration"] <= 7.0)]
        if "text" in train_df.columns:
            text_lens = train_df["text"].astype(str).str.replace(" ", "", regex=False).str.replace("，", "", regex=False).str.replace("。", "", regex=False).str.len()
            dur = train_df["duration"].replace(0, float("nan"))
            cps = (text_lens / dur).fillna(0)
            train_df = train_df[cps >= 1.5]
        train_df = train_df.reset_index(drop=True)
    if "duration" in val_df.columns:
        val_df = val_df[(val_df["duration"] >= 1.5) & (val_df["duration"] <= 7.0)]
        if "text" in val_df.columns:
            text_lens = val_df["text"].astype(str).str.replace(" ", "", regex=False).str.replace("，", "", regex=False).str.replace("。", "", regex=False).str.len()
            dur = val_df["duration"].replace(0, float("nan"))
            cps = (text_lens / dur).fillna(0)
            val_df = val_df[cps >= 1.5]
        val_df = val_df.reset_index(drop=True)
    meta_train = segments_dir / "meta_train_vits.txt"
    meta_val = segments_dir / "meta_val_vits.txt"
    def to_meta(df, out_path):
        lines = []
        for _, row in df.iterrows():
            stem = Path(row["audio_path"]).stem
            text = str(row.get("text", "")).strip()
            lines.append(f"{stem}|female|{text}")
        out_path.write_text("\n".join(lines), encoding="utf-8")
        return len(lines)
    n_train = to_meta(train_df, meta_train)
    n_val = to_meta(val_df, meta_val)
    wavs_dir = root_path / "wavs"
    if wavs_dir.exists() and not wavs_dir.is_symlink():
        pass  # 不覆盖非链接
    elif not wavs_dir.exists() and segments_dir.exists():
        try:
            wavs_dir.symlink_to(segments_dir.resolve())
        except Exception:
            pass
    ASCII_PUNCT = set("!'(),-.:;? \t\n\r")
    all_text = "".join(train_df["text"].astype(str).tolist() + val_df["text"].astype(str).tolist())
    chars_from_data = "".join(sorted(set(c for c in all_text if c not in ASCII_PUNCT)))
    return str(meta_train), str(meta_val), n_train + n_val, chars_from_data


def _run_vits_train_self_contained(data_path, output_dir, epochs):
    """自包含 VITS 训练：仅用本文件与 TTS 库，不调用 vits_moaning_trainer.py。始终从头训练。"""
    from TTS.tts.configs.vits_config import VitsConfig
    from TTS.tts.configs.shared_configs import CharactersConfig
    from TTS.tts.models.vits import Vits, VitsDataset
    # 修复 eval 时 rescue 下标越界：TTS VitsDataset 在 token/audio 超长时用 rescue_item_idx 取下一项，可能超出 len(samples)
    _orig_getitem = VitsDataset.__getitem__
    def _patched_getitem(self, idx):
        n = len(self.samples)
        if n == 0:
            raise IndexError("no samples")
        idx = int(idx) % n
        return _orig_getitem(self, idx)
    VitsDataset.__getitem__ = _patched_getitem
    from TTS.tts.utils.text import tokenizer as _tts_tok_mod
    from TTS.utils.audio import AudioProcessor
    from TTS.tts.datasets import load_tts_samples
    from trainer import Trainer, TrainerArgs
    from TTS.config.shared_configs import BaseDatasetConfig
    from TTS.tts.utils.text.tokenizer import TTSTokenizer

    data_path = Path(data_path).resolve()
    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    audio_dir = "."
    meta_train, meta_val, num_samples, chars_from_data = _prepare_data_tts_for_train(data_path, audio_dir)
    if num_samples == 0 or not chars_from_data:
        print("❌ 无有效训练样本或词表为空，跳过训练")
        return
    print(f"📊 准备完成: train+val={num_samples}, 词表字符数={len(chars_from_data)}")
    config = VitsConfig()
    config.model = "vits"
    config.output_path = str(output_dir)
    config.run_name = "vits_codebuddy_voice"
    config.use_phonemes = False
    config.characters = CharactersConfig(
        pad="<PAD>",
        punctuations="!'(),-.:;? \t\n\r",
        characters=chars_from_data,
        phonemes="",
    )
    config.epochs = int(epochs)
    config.batch_size = 16  # 稍大 batch 利于稳定收敛，显存不足可改为 8
    config.eval_batch_size = min(16, config.batch_size)
    config.save_step = 500
    config.print_step = 50  # 每 50 step 打印一次损失；设为 1 会触发每步验证并卡在 Synthesizing test sentences
    # 关键：关闭训练过程中的验证/测试句合成，否则会一直卡在 "Synthesizing test sentences" 无法进入正式训练
    config.run_eval = False
    config.test_delay_epochs = 9999  # 若日后开启 run_eval，可延后首轮 test_run
    config.num_loader_workers = 4
    config.min_text_len = 1
    config.max_text_len = 50
    config.min_audio_len = 33075
    config.max_audio_len = 154350
    config.dur_loss_alpha = 10.0  # 强制重视时长
    config.mel_loss_alpha = 35.0
    config.datasets = [
        BaseDatasetConfig(
            formatter="ljspeech",
            dataset_name="codebuddy_voice",
            path=str(data_path),
            meta_file_train=f"{audio_dir}/meta_train_vits.txt",
            meta_file_val=f"{audio_dir}/meta_val_vits.txt",
        )
    ]
    config.test_sentences = [["啊"], ["嗯"], ["好的"]]
    config.audio.sample_rate = 22050
    config.audio.fft_size = 1024
    config.audio.win_length = 1024
    config.audio.hop_length = 256
    config.audio.num_mels = 80
    config.audio.mel_fmin = 0
    config.audio.mel_fmax = None
    config.vits = {
        "inter_channels": 192,
        "hidden_channels": 192,
        "filter_channels": 768,
        "n_heads": 2,
        "n_layers": 6,
        "kernel_size": 3,
        "p_dropout": 0.1,
        "resblock": "1",
        "resblock_kernel_sizes": [3, 7, 11],
        "resblock_dilation_sizes": [[1, 3, 5], [1, 3, 5], [1, 3, 5]],
        "upsample_rates": [8, 8, 2, 2],
        "upsample_initial_channel": 512,
        "upsample_kernel_sizes": [16, 16, 4, 4],
        "n_layers_q": 3,
        "use_spectral_norm": False,
    }
    config.train = {"seg_len": 8192, "port": 5000, "c_mel": 35, "c_kl": 1.0, "c_commit": 1.0, "c_reconstruct": 1.0}
    # 学习率：使用 TTS 默认（2e-4）；2000 轮时降至初始的 50%，之后保持该值
    config.lr = 5e-6  # 降低约 40 倍，利于精细收敛
    config.optimizer_params = {"betas": [0.8, 0.99], "eps": 1e-09, "weight_decay": 0.0}
    lr_decay_epoch = 2000
    t_max = lr_decay_epoch
    eta_min = config.lr * 0.5  # 2000 轮时 LR = 50% 初始值 = 1e-4
    config.lr_scheduler = "CosineAnnealingLR"
    config.lr_scheduler_params = {"T_max": t_max, "eta_min": eta_min}
    config.lr_scheduler_gen = "CosineAnnealingLR"
    config.lr_scheduler_gen_params = {"T_max": t_max, "eta_min": eta_min}
    config.lr_scheduler_disc = "CosineAnnealingLR"
    config.lr_scheduler_disc_params = {"T_max": t_max, "eta_min": eta_min}
    config.scheduler_after_epoch = True
    # 梯度裁剪放宽至 1.0，0.3 过紧可能阻碍收敛
    config.grad_clip = [1.0, 1.0]
    config.use_amp = True
    config.pin_memory = True
    train_samples, eval_samples = load_tts_samples(config.datasets, eval_split=True, eval_split_size=0.05)
    if not train_samples:
        print("❌ 未加载到训练样本")
        return
    # 样本很少时保证至少 1 条验证，避免 eval 阶段 "No samples left"
    if not eval_samples and train_samples:
        n_eval = max(1, min(10, len(train_samples) // 10))
        n_eval = min(n_eval, len(train_samples) - 1) if len(train_samples) > 1 else 1
        eval_samples = list(train_samples[-n_eval:])
        train_samples = list(train_samples[:-n_eval])
    if not eval_samples and train_samples:
        eval_samples = list(train_samples[:1])
        train_samples = list(train_samples[1:])
    gpu = torch.cuda.is_available()
    ap = AudioProcessor.init_from_config(config)
    tokenizer, config = TTSTokenizer.init_from_config(config)
    n_chars = tokenizer.characters.num_chars
    required_num_chars = n_chars + 1 if tokenizer.add_blank else n_chars
    config.model_args.num_chars = required_num_chars
    _saved_characters = getattr(config, "characters", None)
    if _saved_characters and required_num_chars != n_chars:
        config.characters = None
    model = Vits(config, ap, tokenizer, speaker_manager=None)
    if _saved_characters:
        config.characters = _saved_characters
    if model.args.num_chars < required_num_chars:
        emb = model.text_encoder.emb
        new_emb = torch.nn.Embedding(
            required_num_chars, emb.embedding_dim,
            padding_idx=getattr(emb, "padding_idx", None),
        ).to(emb.weight.device)
        with torch.no_grad():
            new_emb.weight[: model.args.num_chars].copy_(emb.weight)
        model.text_encoder.emb = new_emb
        model.args.num_chars = required_num_chars
    model.optimize = None
    if gpu:
        model = model.cuda()

    train_args = TrainerArgs(restore_path="", continue_path="", gpu=0 if gpu else None)
    trainer = Trainer(
        train_args,
        config,
        config.output_path,
        model=model,
        train_samples=train_samples,
        eval_samples=eval_samples,
        training_assets={"audio_processor": ap},
        parse_command_line_args=False,
    )
    trainer.config.run_eval = False
    trainer.config.test_delay_epochs = 9999
    
    # 强制重置学习率，避免被 TTS 默认值覆盖
     
    print("🎯 开始训练 VITS（本文件自包含）...")
    print(f"   学习率: {config.lr} → {lr_decay_epoch} 轮时降至 50% 即 {eta_min:.2e}，grad_clip={config.grad_clip}")
    _vits_epoch_timing["train_start"] = time.time()
    _vits_epoch_timing["last_epoch_end"] = None
    try:
        trainer.fit()
    except KeyboardInterrupt:
        print("\n⚠️ 训练被中断")
    except BaseException as e:
        import traceback
        traceback.print_exc()
        raise
    print(f"📁 模型保存: {output_dir}")


# ==================== VITS训练器类 ====================
class VITSTrainer:
    """VITS训练器 - 集成数据清洗功能"""
    
    def __init__(self, config):
        self.config = config
        self.data_path = config.get('data_path', './data')
        self.meta_file = config.get('meta_file', 'metadata.txt')
        self.batch_size = config.get('batch_size', 64)
        self.data_cleaned = False
        self.dataset = None
        self.dataloader = None
        self.model = None
        self.optimizer = None
    
    def clean_data(self, force=False, min_chinese=1, max_samples=0):
        """
        训练前清洗数据
        
        Args:
            force: 强制重新清洗
            min_chinese: 最少中文字符数
            max_samples: 最大样本数
        """
        # 设置清洗配置
        CleanConfig.MIN_CHINESE_CHARS = min_chinese
        CleanConfig.MAX_SAMPLES = max_samples
        
        cleaner = DataCleaner(self.data_path, self.meta_file)
        result = cleaner.run()
        
        # 更新路径为清洗后的路径
        self.data_path = result['output_dir']
        self.meta_file = 'cleaned_metadata.txt'
        self.data_cleaned = True
        
        return result
    
    def _run_build_dataset_from_labels(self, min_chinese=1, max_samples=0):
        """从 labels+downloads 重新构建 segments 与 train/val CSV，并生成 segments/meta_train_vits.txt 供清洗使用（仅用本文件代码）。"""
        if not DATASET_LABELS.is_dir() or not DATASET_DOWNLOADS.is_dir():
            print("⚠️ 未找到 dataset/labels 或 dataset/downloads，跳过从 labels 重新加载（使用已有数据）")
            return False
        print("🔄 从 labels 重新加载数据: build_dataset_from_labels() ...")
        ok = build_dataset_from_labels(
            min_chinese=min_chinese,
            min_duration=1.5,
            max_duration=7.0,
            max_samples=max_samples or 0,
            skip_quality_check=True,
        )
        if not ok:
            print("⚠️ build_dataset_from_labels 未产生数据，继续使用已有数据")
            return False
        # 根据新生成的 train.csv / val.csv 生成 segments/meta_train_vits.txt 供 DataCleaner 使用
        train_csv = ROOT / "train.csv"
        val_csv = ROOT / "val.csv"
        meta_path = ROOT / "segments" / "meta_train_vits.txt"
        meta_path.parent.mkdir(parents=True, exist_ok=True)
        lines = []
        intloop=0
        for csv_path in (train_csv, val_csv):
            if intloop>LOOPMAX:
                break
            intloop+=1
            if not csv_path.exists():
                continue
            df = pd.read_csv(csv_path, dtype={"audio_path": str, "text": str}, encoding="utf-8-sig")
            for _, row in df.iterrows():
                ap, text = row.get("audio_path", ""), row.get("text", "")
                if not ap or pd.isna(ap):
                    continue
                stem = Path(ap).stem
                text = str(text).strip() if not pd.isna(text) else ""
                lines.append(f"{stem}|female|{text}")
        if lines:
            meta_path.write_text("\n".join(lines), encoding="utf-8")
            print(f"✅ 已根据 train/val.csv 生成: {meta_path}（{len(lines)} 条）")
        return True

    def prepare_data(self, skip_cleaning=False, min_chinese=1, max_samples=0):
        """
        准备数据 - 先从 labels 重新加载（若有），再执行清洗步骤
        
        Args:
            skip_cleaning: 跳过清洗
            min_chinese: 最少中文字符数
            max_samples: 最大样本数
        """
        self._run_build_dataset_from_labels(min_chinese=min_chinese, max_samples=max_samples)
        if not skip_cleaning and not self.data_cleaned:
            print("🔄 训练前自动执行数据清洗...")
            result = self.clean_data(min_chinese=min_chinese, max_samples=max_samples)
            print(f"✅ 数据清洗完成，保留 {result['kept']}/{result['total']} 条样本")
            if result['excel_report']:
                print(f"📊 Excel报告: {result['excel_report']}")
        
        # 加载数据集
        print(f"📂 加载数据集: {self.data_path}")
        self.dataset = VITSDataset(self.data_path, self.meta_file)
        n = len(self.dataset)
        if n == 0:
            meta_path = Path(self.data_path) / self.meta_file
            raise ValueError(
                "数据集为空，无法训练。\n"
                f"  当前 data_path = {self.data_path}\n"
                f"  当前 meta_file = {self.meta_file}\n"
                f"  请确认上述路径存在且 {meta_path} 内有有效样本；"
                "若在 IDE 中运行，请将工作目录设为脚本所在目录（lasttraincodebuddy）。"
            )
        if torch:
            self.dataloader = DataLoader(
                self.dataset,
                batch_size=self.batch_size,
                shuffle=True,
                num_workers=8,
                pin_memory=True
            )
            # 训练集 / 验证集划分（约 90% / 10%）
            n = len(self.dataset)
            n_val = max(1, min(n // 10, 500))
            n_train = n - n_val
            train_set, val_set = torch.utils.data.random_split(
                self.dataset, [n_train, n_val],
                generator=torch.Generator().manual_seed(42)
            )
            self.train_loader = DataLoader(
                train_set, batch_size=self.batch_size, shuffle=True,
                num_workers=4, pin_memory=True
            )
            self.val_loader = DataLoader(
                val_set, batch_size=self.batch_size, shuffle=False,
                num_workers=2, pin_memory=True
            )
            self._n_train = n_train
            self._n_val = n_val
            # 写入 TTS 所需 train.csv / val.csv 及 meta（供 vits_moaning_trainer 训练逻辑使用）
            data_path = Path(self.data_path)
            train_indices = self.train_loader.dataset.indices
            val_indices = self.val_loader.dataset.indices
            # 使用文件名即可，因 train() 中会将 moaning.ROOT 设为 data_path（cleaned_audio），路径为 ROOT/文件名
            def _row(i):
                s = self.dataset.samples[i]
                return {"audio_path": Path(s["audio"]).name, "text": s["text"]}
            train_rows = [_row(i) for i in train_indices]
            val_rows = [_row(i) for i in val_indices]
            train_csv = data_path / "train.csv"
            val_csv = data_path / "val.csv"
            pd.DataFrame(train_rows).to_csv(train_csv, index=False, encoding="utf-8-sig")
            pd.DataFrame(val_rows).to_csv(val_csv, index=False, encoding="utf-8-sig")
            for name, rows in [("meta_train_vits.txt", train_rows), ("meta_val_vits.txt", val_rows)]:
                with open(data_path / name, "w", encoding="utf-8") as f:
                    for r in rows:
                        stem = Path(r["audio_path"]).stem
                        f.write(f"{stem}|female|{r['text']}\n")
            print(f"✅ 已写入 TTS 数据: {train_csv}, {val_csv}, meta_train_vits.txt, meta_val_vits.txt")
        else:
            self.train_loader = None
            self.val_loader = None
            self._n_train = n
            self._n_val = 0
        print(f"✅ 数据准备完成: 训练集 {self._n_train} 条, 验证集 {self._n_val} 条")
    
    def train(self, epochs=100, log_path=None):
        """使用本文件内的自包含 VITS 训练逻辑（Coqui TTS）。若 data_path 下已有 train.csv/val.csv 可不执行 prepare_data。始终从头训练。"""
        if not torch:
            print("⚠️ PyTorch 不可用，跳过训练循环")
            return
        data_path = Path(self.data_path).resolve()
        if not (data_path / "train.csv").exists() or not (data_path / "val.csv").exists():
            print("⚠️ 未找到 train.csv/val.csv，请先执行 prepare_data() 或确保数据目录正确")
            return
        print("🆕 从头开始训练：不加载 checkpoint")
        output_dir = VITS_OUTPUT_DIR
        output_dir.mkdir(parents=True, exist_ok=True)
        run_repair=None
        if run_repair is not None:
            print("🔧 步骤0: 音频对齐修复...")
        
            # 检查是否已修复过（避免重复处理）
            repaired_train = ROOT / "repaired_data" / "train.csv"
            repaired_val = ROOT / "repaired_data" / "val.csv"
            
            if not repaired_train.exists():
                # 准备输入JSON（从现有train.csv生成）
                original_train = ROOT / "train.csv"
                if original_train.exists():
                    train_df_original = pd.read_csv(original_train)
                    
                    # 转换为JSON格式
                    input_data = []
                    for _, row in train_df_original.iterrows():
                        input_data.append({
                            "audio_path": str(row["audio_path"]),
                            "text": str(row["text"])
                        })
                    
                    input_json = ROOT / "align_input.json"
                    with open(input_json, 'w', encoding='utf-8') as f:
                        json.dump(input_data, f, ensure_ascii=False, indent=2)
                    
                    print(f"   准备修复 {len(input_data)} 条样本...")
                    
                    # 调用对齐修复
                    try:
                        result = run_repair(
                            input_json=input_json,
                            output_dir=ROOT / "repaired_data",
                            whisper_model_size="base",
                            trim_silence=True,
                            report_excel=ROOT / "align_report.xlsx"
                        )
                        
                        print(f"✅ 对齐修复完成:")
                        print(f"   保留: {result.get('kept', 0)}条")
                        print(f"   ASR替换: {result.get('replaced', 0)}条") 
                        print(f"   需审核: {result.get('review', 0)}条")
                        if result.get('report_excel'):
                            print(f"   报告: {result['report_excel']}")
                    except Exception as e:
                        print(f"⚠️  对齐修复失败: {e}")
                        print("   将使用原始数据继续训练")
            else:
                print("✅ 检测到已修复数据，跳过对齐步骤")
        _run_vits_train_self_contained(data_path, output_dir, int(epochs))
        print("\n✅ 训练完成（本文件自包含逻辑）")


# ==================== 主函数 ====================
def main():
    parser = argparse.ArgumentParser(description='VITS Trainer with Data Cleaning & Excel Export')
    
    # 数据相关（默认使用脚本所在目录下的 segments）
    parser.add_argument('--data_path', type=str, default=DEFAULT_DATA_PATH,
                        help='数据目录路径（默认: 脚本同目录/segments）')
    parser.add_argument('--meta_file', type=str, default=DEFAULT_META_FILE,
                        help='标注文件名（默认 meta_train_vits.txt）')
    
    # 清洗相关
    parser.add_argument('--skip_cleaning', action='store_true',
                        help='跳过数据清洗')
    parser.add_argument('--force_cleaning', action='store_true',
                        help='强制重新清洗')
    parser.add_argument('--min_chinese', type=int, default=1,
                        help='最少中文字符数（默认1）')
    parser.add_argument('--max_samples', type=int, default=0,
                        help='最大样本数（0表示不限制）')
    
    # 训练相关
    parser.add_argument('--batch_size', type=int, default=16,
                        help='批次大小')
    parser.add_argument('--epochs', type=int, default=3000,
                        help='训练轮数（默认3000）')
    
    # 仅清洗模式
    parser.add_argument('--clean_only', action='store_true',
                        help='仅执行数据清洗，不训练')
    # 数据已准备好时跳过准备与检测，直接训练
    parser.add_argument('--train_only', action='store_true',
                        help='跳过数据准备与清洗，直接训练（要求 data_path 下已有 train.csv/val.csv/meta）')
    args = parser.parse_args()

    # 全程日志：train_log_日期时间.txt，与控制台同步输出；训练损失等指标保留 2 位小数显示
    # 同时重定向 stderr，因 Coqui TTS / logging 默认写 stderr，否则日志文件会缺少大量训练输出
    log_stem = f"train_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    log_path = SCRIPT_DIR / f"{log_stem}.txt"
    tee = TeeLogger(sys.stdout, log_path)
    _stdout_orig, _stderr_orig = sys.stdout, sys.stderr
    inner = LossFormatStream(tee)
    epoch_stream = EpochTimingStream(inner)
    sys.stdout = epoch_stream
    sys.stderr = epoch_stream
    try:
        print(f"📄 训练日志已写入: {log_path}")
        print("=" * 60)

        # 配置
        config = {
            'data_path': args.data_path,
            'meta_file': args.meta_file,
            'batch_size': args.batch_size,
        }

        # 初始化训练器
        trainer = VITSTrainer(config)

        # 仅清洗模式
        if args.clean_only:
            print("🧹 仅执行数据清洗模式")
            result = trainer.clean_data(
                force=args.force_cleaning,
                min_chinese=args.min_chinese,
                max_samples=args.max_samples
            )
            print(f"\n清洗结果:")
            print(f"  总样本: {result['total']}")
            print(f"  保留: {result['kept']}")
            print(f"  Excel报告: {result['excel_report']}")
            return

        # 数据已准备则跳过准备与检测，直接进入训练
        data_ready_path = _get_train_data_path(args.data_path)
        if args.train_only or data_ready_path is not None:
            if data_ready_path is None:
                data_ready_path = Path(args.data_path).resolve()
                if not _is_tts_data_ready(data_ready_path):
                    print("⚠️ --train_only 要求 data_path 下已有 train.csv、val.csv、meta_train_vits.txt")
                    return
            trainer.data_path = str(data_ready_path)
            trainer.meta_file = "meta_train_vits.txt"
            print(f"✅ 检测到数据已准备，跳过数据准备与清洗，直接训练: {trainer.data_path}")
        else:
            # 完整流程：从 labels 构建、清洗、再训练
            trainer.prepare_data(
                skip_cleaning=args.skip_cleaning,
                min_chinese=args.min_chinese,
                max_samples=args.max_samples
            )

        # 开始训练（始终从头训练）
        trainer.train(epochs=args.epochs, log_path=log_path)

        print("=" * 60)
        print(f"📄 完整日志已保存: {log_path}")
    finally:
        sys.stdout.flush()
        sys.stderr.flush()
        sys.stdout = _stdout_orig
        sys.stderr = _stderr_orig
        tee.close()


if __name__ == '__main__':
    main()
